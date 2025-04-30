import gc
from pathlib import Path

import openpyxl
import pytest

from otlmow_template.SubsetTemplateCreator import SubsetTemplateCreator

current_dir = Path(__file__).parent
model_directory_path = Path(__file__).parent.parent / 'TestModel'


def test_subset_with_AllCasesTestClass_0_records_relations():
    subset_tool = SubsetTemplateCreator()
    excel_path = current_dir / 'OTL_AllCasesTestClass_relations_1.xlsx'
    subset_tool.generate_template_from_subset(subset_path=current_dir / 'OTL_AllCasesTestClass.db',
                                              template_file_path=excel_path, model_directory=model_directory_path,
                                              dummy_data_rows=0, ignore_relations=False)
    assert excel_path.exists()

    book = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    header_row_list = []
    bevestiging_sheet_data = []

    for sheet in book.worksheets:
        sheet_name = sheet.title
        if sheet_name == 'onderdeel#AllCasesTestClass':
            for row in sheet.rows:
                header_row_list = [cell.value for cell in row]
                break
        elif sheet_name == 'onderdeel#Bevestiging':
            for row in sheet.rows:
                bevestiging_sheet_data.append([cell.value for cell in row])
    book.close()

    union_headers = [header for header in header_row_list if header.startswith('testUnionType')]
    header_row_list = [header for header in header_row_list if not header.startswith('testUnionType')]

    assert bevestiging_sheet_data == [['typeURI',
      'assetId.identificator',
      'assetId.toegekendDoor',
      'bron.typeURI',
      'bronAssetId.identificator',
      'bronAssetId.toegekendDoor',
      'doel.typeURI',
      'doelAssetId.identificator',
      'doelAssetId.toegekendDoor']]

    assert header_row_list == ['typeURI', 'assetId.identificator', 'assetId.toegekendDoor', 'bestekPostNummer[]',
                               'datumOprichtingObject', 'geometry', 'isActief', 'notitie',
                               'standaardBestekPostNummer[]',
                               'testBooleanField', 'testComplexType.testBooleanField',
                               'testComplexType.testComplexType2.testKwantWrd',
                               'testComplexType.testComplexType2.testStringField',
                               'testComplexType.testComplexType2MetKard[].testKwantWrd',
                               'testComplexType.testComplexType2MetKard[].testStringField',
                               'testComplexType.testKwantWrd',
                               'testComplexType.testKwantWrdMetKard[]',
                               'testComplexType.testStringField',
                               'testComplexType.testStringFieldMetKard[]',
                               'testComplexTypeMetKard[].testBooleanField',
                               'testComplexTypeMetKard[].testComplexType2.testKwantWrd',
                               'testComplexTypeMetKard[].testComplexType2.testStringField',
                               'testComplexTypeMetKard[].testKwantWrd',
                               'testComplexTypeMetKard[].testStringField',
                               'testDateField', 'testDateTimeField', 'testDecimalField',
                               'testDecimalFieldMetKard[]', 'testEenvoudigType', 'testEenvoudigTypeMetKard[]',
                               'testIntegerField', 'testIntegerFieldMetKard[]', 'testKeuzelijst',
                               'testKeuzelijstMetKard[]', 'testKwantWrd', 'testKwantWrdMetKard[]',
                               'testStringField', 'testStringFieldMetKard[]', 'testTimeField', 'theoretischeLevensduur',
                               'toestand']

    assert union_headers[0].startswith('testUnionType.')
    assert union_headers[1].startswith('testUnionTypeMetKard[].')

    gc.collect()

    excel_path.unlink()



def test_subset_with_camera_steun_0_records_relations():
    subset_tool = SubsetTemplateCreator()
    excel_path = current_dir / 'OTL_camera_steun_relations_1.xlsx'
    subset_tool.generate_template_from_subset(subset_path=current_dir / 'camera_steun_2.14.db',
                                              template_file_path=excel_path,
                                              dummy_data_rows=0, ignore_relations=False)
    assert excel_path.exists()

    book = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    header_row_list = []
    bevestiging_sheet_data = []

    for sheet in book.worksheets:
        sheet_name = sheet.title
        if sheet_name == 'onderdeel#Camera':
            for row in sheet.rows:
                header_row_list = [cell.value for cell in row]
                break
        elif sheet_name == 'onderdeel#Bevestiging':
            for row in sheet.rows:
                bevestiging_sheet_data.append([cell.value for cell in row])
    book.close()

    assert bevestiging_sheet_data == [['typeURI',
      'assetId.identificator',
      'assetId.toegekendDoor',
      'bron.typeURI',
      'bronAssetId.identificator',
      'bronAssetId.toegekendDoor',
      'doel.typeURI',
      'doelAssetId.identificator',
      'doelAssetId.toegekendDoor']]

    assert header_row_list == ['typeURI',
        'assetId.identificator',
        'assetId.toegekendDoor',
        'beeldverwerkingsinstelling[].configBestand.bestandsnaam',
        'beeldverwerkingsinstelling[].configBestand.mimeType',
        'beeldverwerkingsinstelling[].configBestand.omschrijving',
        'beeldverwerkingsinstelling[].configBestand.opmaakdatum',
        'beeldverwerkingsinstelling[].configBestand.uri',
        'beeldverwerkingsinstelling[].typeBeeldverwerking',
        'bestekPostNummer[]',
        'datumOprichtingObject',
        'dnsNaam',
        'geometry',
        'heeftFlits',
        'ipAdres',
        'isActief',
        'isPtz',
        'merk',
        'modelnaam',
        'naam',
        'notitie',
        'opstelhoogte',
        'opstelwijze',
        'rijrichting',
        'serienummer',
        'spectrum',
        'standaardBestekPostNummer[]',
        'technischeFiche[].bestandsnaam',
        'technischeFiche[].mimeType',
        'technischeFiche[].omschrijving',
        'technischeFiche[].opmaakdatum',
        'technischeFiche[].uri',
        'theoretischeLevensduur',
        'toestand']

    gc.collect()

    excel_path.unlink()

def test_subset_with_camera_steun_relations_1_selected_class():
    subset_tool = SubsetTemplateCreator()
    excel_path = current_dir / 'OTL_camera_steun_relations_2.xlsx'
    subset_tool.generate_template_from_subset(
        subset_path=current_dir / 'camera_steun_2.14.db', template_file_path=excel_path,
        class_uris_filter=['https://wegenenverkeer.data.vlaanderen.be/ns/onderdeel#Camera'],
        dummy_data_rows=1, ignore_relations=False)
    assert excel_path.exists()

    book = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    header_row_list = []
    bevestiging_sheet_data = []

    for sheet in book.worksheets:
        sheet_name = sheet.title
        if sheet_name == 'onderdeel#Camera':
            for row in sheet.rows:
                header_row_list = [cell.value for cell in row]
                break
        elif sheet_name == 'onderdeel#Bevestiging':
            for row in sheet.rows:
                bevestiging_sheet_data.append([cell.value for cell in row])
    book.close()

    assert bevestiging_sheet_data == []

    assert header_row_list == ['typeURI',
        'assetId.identificator',
        'assetId.toegekendDoor',
        'beeldverwerkingsinstelling[].configBestand.bestandsnaam',
        'beeldverwerkingsinstelling[].configBestand.mimeType',
        'beeldverwerkingsinstelling[].configBestand.omschrijving',
        'beeldverwerkingsinstelling[].configBestand.opmaakdatum',
        'beeldverwerkingsinstelling[].configBestand.uri',
        'beeldverwerkingsinstelling[].typeBeeldverwerking',
        'bestekPostNummer[]',
        'datumOprichtingObject',
        'dnsNaam',
        'geometry',
        'heeftFlits',
        'ipAdres',
        'isActief',
        'isPtz',
        'merk',
        'modelnaam',
        'naam',
        'notitie',
        'opstelhoogte',
        'opstelwijze',
        'rijrichting',
        'serienummer',
        'spectrum',
        'standaardBestekPostNummer[]',
        'technischeFiche[].bestandsnaam',
        'technischeFiche[].mimeType',
        'technischeFiche[].omschrijving',
        'technischeFiche[].opmaakdatum',
        'technischeFiche[].uri',
        'theoretischeLevensduur',
        'toestand']

    gc.collect()

    excel_path.unlink()
