import gc
from pathlib import Path

import openpyxl
import pytest

from otlmow_template.SubsetTemplateCreator import SubsetTemplateCreator

current_dir = Path(__file__).parent
model_directory_path = Path(__file__).parent.parent / 'TestModel'


def test_subset_with_AllCasesTestClass_0_records_relations():
    subset_tool = SubsetTemplateCreator()
    excel_path = current_dir / 'OTL_AllCasesTestClass_relations_1.xlsx'
    subset_tool.generate_template_from_subset(subset_path=current_dir / 'OTL_AllCasesTestClass.db',
                                              template_file_path=excel_path, model_directory=model_directory_path,
                                              dummy_data_rows=0, ignore_relations=False)
    assert excel_path.exists()

    book = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    header_row_list = []
    bevestiging_sheet_data = []

    for sheet in book.worksheets:
        sheet_name = sheet.title
        if sheet_name == 'onderdeel#AllCasesTestClass':
            for row in sheet.rows:
                header_row_list = [cell.value for cell in row]
                break
        elif sheet_name == 'onderdeel#Bevestiging':
            for row in sheet.rows:
                bevestiging_sheet_data.append([cell.value for cell in row])
    book.close()

    union_headers = [header for header in header_row_list if header.startswith('testUnionType')]
    header_row_list = [header for header in header_row_list if not header.startswith('testUnionType')]

    assert bevestiging_sheet_data == [['typeURI',
      'assetId.identificator',
      'assetId.toegekendDoor',
      'bron.typeURI',
      'bronAssetId.identificator',
      'bronAssetId.toegekendDoor',
      'doel.typeURI',
      'doelAssetId.identificator',
      'doelAssetId.toegekendDoor']]

    assert header_row_list == ['typeURI', 'assetId.identificator', 'assetId.toegekendDoor', 'bestekPostNummer[]',
                               'datumOprichtingObject', 'geometry', 'isActief', 'notitie',
                               'standaardBestekPostNummer[]',
                               'testBooleanField', 'testComplexType.testBooleanField',
                               'testComplexType.testComplexType2.testKwantWrd',
                               'testComplexType.testComplexType2.testStringField',
                               'testComplexType.testComplexType2MetKard[].testKwantWrd',
                               'testComplexType.testComplexType2MetKard[].testStringField',
                               'testComplexType.testKwantWrd',
                               'testComplexType.testKwantWrdMetKard[]',
                               'testComplexType.testStringField',
                               'testComplexType.testStringFieldMetKard[]',
                               'testComplexTypeMetKard[].testBooleanField',
                               'testComplexTypeMetKard[].testComplexType2.testKwantWrd',
                               'testComplexTypeMetKard[].testComplexType2.testStringField',
                               'testComplexTypeMetKard[].testKwantWrd',
                               'testComplexTypeMetKard[].testStringField',
                               'testDateField', 'testDateTimeField', 'testDecimalField',
                               'testDecimalFieldMetKard[]', 'testEenvoudigType', 'testEenvoudigTypeMetKard[]',
                               'testIntegerField', 'testIntegerFieldMetKard[]', 'testKeuzelijst',
                               'testKeuzelijstMetKard[]', 'testKwantWrd', 'testKwantWrdMetKard[]',
                               'testStringField', 'testStringFieldMetKard[]', 'testTimeField', 'theoretischeLevensduur',
                               'toestand']

    assert union_headers[0].startswith('testUnionType.')
    assert union_headers[1].startswith('testUnionTypeMetKard[].')

    gc.collect()

    excel_path.unlink()



def test_subset_slagboom_0_records_relations():
    subset_tool = SubsetTemplateCreator()
    excel_path = current_dir / 'OTL_slagboom_relations_0.xlsx'
    subset_tool.generate_template_from_subset(subset_path=current_dir / 'voorbeeld-slagboom.db',
                                              template_file_path=excel_path, add_attribute_info=True,
                                              add_geo_artefact=True, dummy_data_rows=0, ignore_relations=False)
    assert excel_path.exists()

    book = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    arm_header_row_list = []
    arm_data_row_list = []
    bevestiging_header_row_list = []
    bevestiging_data_row_list = []

    for sheet in book.worksheets:
        sheet_name = sheet.title
        if sheet_name == 'onderdeel#Bevestiging':
            bevestiging_header_row_list.extend(
                [cell.value for cell in row]
                for row in sheet.iter_rows(min_row=2, max_row=2)
            )
            bevestiging_data_row_list.extend(
                [cell.value for cell in row]
                for row in sheet.iter_rows(min_row=3, max_row=10)
            )
        elif sheet_name == 'onderdeel#Slagboomarm':
            arm_header_row_list.extend(
                [cell.value for cell in row]
                for row in sheet.iter_rows(min_row=2, max_row=2)
            )
            arm_data_row_list.extend(
                [cell.value for cell in row]
                for row in sheet.iter_rows(min_row=3, max_row=10)
            )
    book.close()

    assert bevestiging_header_row_list == [['typeURI',
      'assetId.identificator',
      'assetId.toegekendDoor',
      'bron.typeURI',
      'bronAssetId.identificator',
      'bronAssetId.toegekendDoor',
      'doel.typeURI',
      'doelAssetId.identificator',
      'doelAssetId.toegekendDoor']]

    assert not bevestiging_data_row_list

    assert arm_header_row_list == [['typeURI',
        'assetId.identificator',
        'assetId.toegekendDoor',
        'bestekPostNummer[]',
        'datumOprichtingObject',
        'geometry',
        'isActief',
        'lengteBoom',
        'notitie',
        'standaardBestekPostNummer[]',
        'technischeFiche.bestandsnaam',
        'technischeFiche.mimeType',
        'technischeFiche.omschrijving',
        'technischeFiche.opmaakdatum',
        'technischeFiche.uri',
        'theoretischeLevensduur',
        'toestand']]

    gc.collect()

    excel_path.unlink()

def test_subset_slagboom_relations_1_selected_class():
    subset_tool = SubsetTemplateCreator()
    excel_path = current_dir / 'OTL_slagboom_relations_1.xlsx'
    subset_tool.generate_template_from_subset(
        subset_path=current_dir / 'voorbeeld-slagboom.db', add_attribute_info=True, template_file_path=excel_path,
        class_uris_filter=['https://wegenenverkeer.data.vlaanderen.be/ns/onderdeel#Slagboomarm'],
        dummy_data_rows=1, ignore_relations=False)
    assert excel_path.exists()

    book = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    arm_header_row_list = []
    arm_data_row_list = []
    bevestiging_header_row_list = []
    bevestiging_data_row_list = []

    for sheet in book.worksheets:
        sheet_name = sheet.title
        if sheet_name == 'onderdeel#Bevestiging':
            bevestiging_header_row_list.extend(
                [cell.value for cell in row]
                for row in sheet.iter_rows(min_row=2, max_row=2)
            )
            bevestiging_data_row_list.extend(
                [cell.value for cell in row]
                for row in sheet.iter_rows(min_row=3, max_row=3)
            )
        elif sheet_name == 'onderdeel#Slagboomarm':
            arm_header_row_list.extend(
                [cell.value for cell in row]
                for row in sheet.iter_rows(min_row=2, max_row=2)
            )
            arm_data_row_list.extend(
                [cell.value for cell in row]
                for row in sheet.iter_rows(min_row=3, max_row=3)
            )
    book.close()

    assert bevestiging_header_row_list == []

    assert len(bevestiging_data_row_list) == 0

    assert arm_data_row_list != []

    assert arm_header_row_list == [['typeURI',
                                    'assetId.identificator',
                                    'assetId.toegekendDoor',
                                    'bestekPostNummer[]',
                                    'datumOprichtingObject',
                                    'geometry',
                                    'isActief',
                                    'lengteBoom',
                                    'notitie',
                                    'standaardBestekPostNummer[]',
                                    'technischeFiche.bestandsnaam',
                                    'technischeFiche.mimeType',
                                    'technischeFiche.omschrijving',
                                    'technischeFiche.opmaakdatum',
                                    'technischeFiche.uri',
                                    'theoretischeLevensduur',
                                    'toestand']]

    gc.collect()

    excel_path.unlink()



def test_subset_slagboom_relations_2_selected_classes_2_rows():
    subset_tool = SubsetTemplateCreator()
    excel_path = current_dir / 'OTL_slagboom_relations_2.xlsx'
    subset_tool.generate_template_from_subset(
        subset_path=current_dir / 'voorbeeld-slagboom.db', add_attribute_info=True, template_file_path=excel_path,
        class_uris_filter=['https://wegenenverkeer.data.vlaanderen.be/ns/onderdeel#Slagboomarm',
                           'https://wegenenverkeer.data.vlaanderen.be/ns/onderdeel#SlagboomarmVerlichting'],
        dummy_data_rows=2, ignore_relations=False)
    assert excel_path.exists()

    book = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    arm_header_row_list = []
    arm_data_row_list = []
    bevestiging_header_row_list = []
    bevestiging_data_row_list = []

    for sheet in book.worksheets:
        sheet_name = sheet.title
        if sheet_name == 'onderdeel#Bevestiging':
            bevestiging_header_row_list.extend(
                [cell.value for cell in row]
                for row in sheet.iter_rows(min_row=2, max_row=2)
            )
            bevestiging_data_row_list.extend(
                [cell.value for cell in row]
                for row in sheet.iter_rows(min_row=3, max_row=10)
            )
        elif sheet_name == 'onderdeel#Slagboomarm':
            arm_header_row_list.extend(
                [cell.value for cell in row]
                for row in sheet.iter_rows(min_row=2, max_row=2)
            )
            arm_data_row_list.extend(
                [cell.value for cell in row]
                for row in sheet.iter_rows(min_row=3, max_row=10)
            )
    book.close()

    assert bevestiging_header_row_list != []

    assert len(bevestiging_data_row_list) == 2

    assert arm_data_row_list != []

    assert arm_header_row_list == [['typeURI',
                                    'assetId.identificator',
                                    'assetId.toegekendDoor',
                                    'bestekPostNummer[]',
                                    'datumOprichtingObject',
                                    'geometry',
                                    'isActief',
                                    'lengteBoom',
                                    'notitie',
                                    'standaardBestekPostNummer[]',
                                    'technischeFiche.bestandsnaam',
                                    'technischeFiche.mimeType',
                                    'technischeFiche.omschrijving',
                                    'technischeFiche.opmaakdatum',
                                    'technischeFiche.uri',
                                    'theoretischeLevensduur',
                                    'toestand']]

    gc.collect()

    excel_path.unlink()


def test_subset_slagboom_relations_2_non_related_selected_classes_2_rows():
    subset_tool = SubsetTemplateCreator()
    excel_path = current_dir / 'OTL_slagboom_relations_3.xlsx'
    subset_tool.generate_template_from_subset(
        subset_path=current_dir / 'voorbeeld-slagboom.db', add_attribute_info=True, template_file_path=excel_path,
        class_uris_filter=['https://wegenenverkeer.data.vlaanderen.be/ns/onderdeel#Slagboomarm',
                           'https://wegenenverkeer.data.vlaanderen.be/ns/installatie#Kokerafsluiting'],
        dummy_data_rows=2, ignore_relations=False)
    assert excel_path.exists()

    book = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    arm_header_row_list = []
    arm_data_row_list = []
    bevestiging_header_row_list = []
    bevestiging_data_row_list = []

    for sheet in book.worksheets:
        sheet_name = sheet.title
        if sheet_name == 'onderdeel#Bevestiging':
            bevestiging_header_row_list.extend(
                [cell.value for cell in row]
                for row in sheet.iter_rows(min_row=2, max_row=2)
            )
            bevestiging_data_row_list.extend(
                [cell.value for cell in row]
                for row in sheet.iter_rows(min_row=3, max_row=10)
            )
        elif sheet_name == 'onderdeel#Slagboomarm':
            arm_header_row_list.extend(
                [cell.value for cell in row]
                for row in sheet.iter_rows(min_row=2, max_row=2)
            )
            arm_data_row_list.extend(
                [cell.value for cell in row]
                for row in sheet.iter_rows(min_row=3, max_row=10)
            )
    book.close()

    assert bevestiging_header_row_list == []

    assert len(bevestiging_data_row_list) == 0

    assert arm_data_row_list != []

    assert arm_header_row_list == [['typeURI',
                                    'assetId.identificator',
                                    'assetId.toegekendDoor',
                                    'bestekPostNummer[]',
                                    'datumOprichtingObject',
                                    'geometry',
                                    'isActief',
                                    'lengteBoom',
                                    'notitie',
                                    'standaardBestekPostNummer[]',
                                    'technischeFiche.bestandsnaam',
                                    'technischeFiche.mimeType',
                                    'technischeFiche.omschrijving',
                                    'technischeFiche.opmaakdatum',
                                    'technischeFiche.uri',
                                    'theoretischeLevensduur',
                                    'toestand']]

    gc.collect()

    excel_path.unlink()


